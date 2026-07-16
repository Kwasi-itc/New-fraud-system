from __future__ import annotations

import csv
import json
import re
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook


@dataclass(frozen=True)
class DedupeStats:
    source_rows: int
    unique_keys: int
    duplicate_rows: int
    conflicting_keys: int
    exact_only_duplicate_keys: int


@dataclass(frozen=True)
class ReferenceRecords:
    records: list[dict[str, Any]]
    stats: DedupeStats


@dataclass(frozen=True)
class StaffLists:
    staff_numbers: tuple[str, ...]
    emails: tuple[str, ...]
    msisdns: tuple[str, ...]
    source_rows: int


@dataclass(frozen=True)
class MerchantWatchlist:
    names: tuple[str, ...]
    source_rows: int
    duplicate_names: int
    missing_names: int


def load_merchants(files: list[Path]) -> ReferenceRecords:
    return _load_and_dedupe(
        files,
        key_name="id",
        rank=lambda row, file_index, row_index: (
            _timestamp_rank(row.get("updatedAt")),
            _timestamp_rank(row.get("createdAt")),
            file_index,
            row_index,
        ),
        mapper=_merchant_record,
    )


def load_merchant_products(files: list[Path]) -> ReferenceRecords:
    return _load_and_dedupe(
        files,
        key_name="merchantProductId",
        rank=lambda _row, file_index, row_index: (file_index, row_index),
        mapper=_merchant_product_record,
    )


def load_staff_lists(path: Path) -> StaffLists:
    staff_numbers: set[str] = set()
    emails: set[str] = set()
    msisdns: set[str] = set()
    source_rows = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"STAFF_NO", "EMAIL", "MSISDN"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"{path} is missing required columns: {sorted(missing)}")
        for row in reader:
            source_rows += 1
            staff_number = _string(row.get("STAFF_NO"))
            email = _string(row.get("EMAIL"))
            if staff_number:
                staff_numbers.add(staff_number.upper())
            if email:
                emails.add(email.lower())
            for digits in re.findall(r"\d+", row.get("MSISDN") or ""):
                normalized = _normalize_msisdn(digits)
                if normalized:
                    msisdns.add(normalized)
    return StaffLists(tuple(sorted(staff_numbers)), tuple(sorted(emails)), tuple(sorted(msisdns)), source_rows)


def load_merchant_watchlist(path: Path | None) -> MerchantWatchlist | None:
    if path is None:
        return None
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        normalized_headers = [str(value).strip() if value is not None else "" for value in (headers or ())]
        if "Name" not in normalized_headers:
            raise ValueError(f"{path} must contain a Name column")
        name_index = normalized_headers.index("Name")
        names: set[str] = set()
        source_rows = 0
        missing_names = 0
        duplicate_names = 0
        for row in rows:
            if not any(value is not None and str(value).strip() for value in row):
                continue
            source_rows += 1
            name = _normalize_name(row[name_index] if name_index < len(row) else None)
            if not name:
                missing_names += 1
            elif name in names:
                duplicate_names += 1
            else:
                names.add(name)
        return MerchantWatchlist(tuple(sorted(names)), source_rows, duplicate_names, missing_names)
    finally:
        workbook.close()


def _load_and_dedupe(
    files: list[Path],
    *,
    key_name: str,
    rank: Callable[[dict[str, Any], int, int], tuple[Any, ...]],
    mapper: Callable[[dict[str, Any]], dict[str, Any]],
) -> ReferenceRecords:
    chosen: dict[str, tuple[tuple[Any, ...], dict[str, Any]]] = {}
    variants: dict[str, set[str]] = defaultdict(set)
    counts: dict[str, int] = defaultdict(int)
    source_rows = 0
    for file_index, path in enumerate(files):
        raw = json.loads(path.read_text(encoding="utf-8-sig"))
        if not isinstance(raw, list):
            raise ValueError(f"{path} must contain a JSON array")
        for row_index, row in enumerate(raw):
            source_rows += 1
            if not isinstance(row, dict):
                raise ValueError(f"{path} row {row_index + 1} must be an object")
            key = _string(row.get(key_name))
            if not key:
                raise ValueError(f"{path} row {row_index + 1} has no {key_name}")
            canonical = json.dumps(row, sort_keys=True, separators=(",", ":"), default=str)
            variants[key].add(canonical)
            counts[key] += 1
            candidate_rank = rank(row, file_index, row_index)
            previous = chosen.get(key)
            if previous is None or candidate_rank >= previous[0]:
                chosen[key] = (candidate_rank, row)

    records = [mapper(chosen[key][1]) for key in sorted(chosen)]
    duplicate_keys = [key for key, count in counts.items() if count > 1]
    stats = DedupeStats(
        source_rows=source_rows,
        unique_keys=len(chosen),
        duplicate_rows=source_rows - len(chosen),
        conflicting_keys=sum(1 for key in duplicate_keys if len(variants[key]) > 1),
        exact_only_duplicate_keys=sum(1 for key in duplicate_keys if len(variants[key]) == 1),
    )
    return ReferenceRecords(records, stats)


def _merchant_record(row: dict[str, Any]) -> dict[str, Any]:
    merchant_id = str(row["id"])
    return {
        "object_id": merchant_id,
        "merchant_id": merchant_id,
        "company_name": _string(row.get("companyName")),
        "company_name_normalized": _normalize_name(row.get("companyName")),
        "code": _string(row.get("code")),
        "trade_name": _string(row.get("tradeName")),
        "alias": _string(row.get("alias")),
        "country": _string(row.get("country")),
        "company_type": _string(row.get("typeOfCompany")),
        "status": _string(row.get("status")),
        "company_registration_number": _string(row.get("companyRegistrationNumber")),
        "tax_identification_number": _string(row.get("taxIdentificationNumber")),
        "created_at": _timestamp_or_none(row.get("createdAt")),
        "updated_at_source": _timestamp_or_none(row.get("updatedAt")),
        "date_of_incorporation": _string(row.get("dateOfIncorporation")),
        "date_of_commencement": _string(row.get("dateOfCommencement")),
    }


def _merchant_product_record(row: dict[str, Any]) -> dict[str, Any]:
    merchant_product_id = str(row["merchantProductId"])
    return {
        "object_id": merchant_product_id,
        "merchant_product_id": merchant_product_id,
        "catalog_product_id": _string(row.get("productId")),
        "product_name": _string(row.get("productName")),
        "name": _string(row.get("name")),
        "can_settle": _string(row.get("canSettle")),
        "merchant_id": _string(row.get("merchantId")),
        "status": _string(row.get("status")),
        "description": _string(row.get("description")),
    }


def _timestamp_rank(value: Any) -> float:
    parsed = _parse_timestamp(value)
    return parsed.timestamp() if parsed else float("-inf")


def _timestamp_or_none(value: Any) -> str | None:
    parsed = _parse_timestamp(value)
    return parsed.isoformat().replace("+00:00", "Z") if parsed else None


def _parse_timestamp(value: Any) -> datetime | None:
    text = _string(value)
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _string(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def _normalize_name(value: Any) -> str | None:
    result = _string(value)
    return re.sub(r"\s+", " ", result).casefold() if result else None


def _normalize_msisdn(value: str) -> str | None:
    digits = re.sub(r"\D", "", value)
    if not digits:
        return None
    if digits.startswith("0") and len(digits) == 10:
        return "233" + digits[1:]
    if digits.startswith("233"):
        return digits
    return digits
